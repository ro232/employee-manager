import os
import io
import shutil
import difflib
from datetime import date, datetime, timedelta
from functools import wraps
from collections import defaultdict

from flask import (
    Flask, render_template, request, redirect, url_for,
    flash, jsonify, send_file, session, g, abort
)
from flask_login import (
    LoginManager, login_user, logout_user, login_required, current_user
)
from config import Config
from models import (
    db, User, Angajat, Hotel, Pontaj, Firma, ContractAngajat,
    AuditLog, Notification, DuplicateExclusion, Planificare
)
from import_excel import (
    parse_excel_file, import_entries, create_angajat_from_name,
    process_zip_file, find_angajat_by_name
)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ---------------------------------------------------------------------------
# Role-based access decorator
# ---------------------------------------------------------------------------
def require_role(*allowed_roles):
    """Decorator that checks current_user.role against allowed roles.
    Usage:  @require_role("admin", "editor")
    """
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not current_user.is_authenticated:
                return redirect(url_for("login"))
            if current_user.role not in allowed_roles:
                flash("Nu ai permisiunea necesara pentru aceasta actiune.", "danger")
                return redirect(url_for("index"))
            return f(*args, **kwargs)
        return decorated_function
    return decorator


def create_app():
    app = Flask(__name__)
    app.config.from_object(Config)

    os.makedirs(app.config.get("UPLOAD_FOLDER", "uploads"), exist_ok=True)
    os.makedirs(os.path.join(app.instance_path), exist_ok=True)
    os.makedirs(os.path.join(app.instance_path, "backups"), exist_ok=True)

    db.init_app(app)
    login_manager = LoginManager()
    login_manager.init_app(app)
    login_manager.login_view = "login"
    login_manager.login_message = "Trebuie sa te autentifici."

    @login_manager.user_loader
    def load_user(user_id):
        return User.query.get(int(user_id))

    with app.app_context():
        db.create_all()
        seed_default_data()

    @app.context_processor
    def inject_globals():
        notif_count = 0
        dark_mode = False
        if current_user.is_authenticated:
            notif_count = Notification.query.filter_by(
                user_id=current_user.id, citit=False
            ).count()
            dark_mode = current_user.dark_mode
        return dict(notif_count=notif_count, dark_mode=dark_mode)

    register_routes(app)
    return app


def seed_default_data():
    if not Firma.query.first():
        firme = [
            Firma(cod="D", nume="Deltha"),
            Firma(cod="E", nume="E-thema"),
            Firma(cod="None", nume="Fara firma"),
        ]
        db.session.add_all(firme)
        db.session.commit()
    if not User.query.first():
        admin = User(username="admin", nume_complet="Administrator",
                     is_admin=True, role="admin")
        admin.set_password("admin")
        db.session.add(admin)
        db.session.commit()
    else:
        # Ensure existing admin user has role="admin"
        admin_user = User.query.filter_by(username="admin").first()
        if admin_user and admin_user.role != "admin":
            admin_user.role = "admin"
            db.session.commit()


def log_audit(actiune, entitate, entitate_id=None, detalii=None):
    user_id = current_user.id if current_user.is_authenticated else None
    entry = AuditLog(
        user_id=user_id, actiune=actiune, entitate=entitate,
        entitate_id=entitate_id, detalii=detalii
    )
    db.session.add(entry)


def add_notification(mesaj, tip="info", link=None, user_id=None):
    if user_id is None and current_user.is_authenticated:
        user_id = current_user.id
    # Notify all users if user_id is None
    if user_id:
        notif = Notification(user_id=user_id, mesaj=mesaj, tip=tip, link=link)
        db.session.add(notif)
    else:
        for u in User.query.all():
            notif = Notification(user_id=u.id, mesaj=mesaj, tip=tip, link=link)
            db.session.add(notif)


def find_similar_names(name, threshold=0.8):
    """Find existing employees with similar names."""
    all_angajati = Angajat.query.filter_by(activ=True).all()
    similar = []
    for a in all_angajati:
        ratio = difflib.SequenceMatcher(None, name.lower(), a.nume_complet.lower()).ratio()
        if ratio >= threshold and name.lower() != a.nume_complet.lower():
            similar.append({"angajat": a, "similarity": round(ratio * 100)})
    return sorted(similar, key=lambda x: x["similarity"], reverse=True)


def build_excel_report(pontaje, filters, grupare):
    """Build Excel workbook from report data."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Raport"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    ws.merge_cells("A1:G1")
    ws["A1"] = f"Raport Pontaj - {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws["A1"].font = Font(bold=True, size=14)

    row = 2
    filter_parts = []
    if filters.get("data_start"):
        filter_parts.append(f"De la: {filters['data_start']}")
    if filters.get("data_end"):
        filter_parts.append(f"Pana la: {filters['data_end']}")
    if filters.get("angajat_id"):
        a = Angajat.query.get(int(filters["angajat_id"]))
        if a:
            filter_parts.append(f"Angajat: {a.nume_complet}")
    if filters.get("hotel_id"):
        h = Hotel.query.get(int(filters["hotel_id"]))
        if h:
            filter_parts.append(f"Hotel: {h.nume}")
    if filters.get("firma_cod"):
        filter_parts.append(f"Firma: {filters['firma_cod']}")
    if filter_parts:
        ws.merge_cells(f"A{row}:G{row}")
        ws[f"A{row}"] = "Filtre: " + " | ".join(filter_parts)
        row += 1
    row += 1

    if grupare == "detaliat":
        headers = ["Data", "Angajat", "Hotel", "Firma", "Ore", "Cost (RON)", "Saptamana"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

        total_cost = 0
        for p in pontaje:
            row += 1
            ws.cell(row=row, column=1, value=p.data.strftime("%d.%m.%Y")).border = thin_border
            ws.cell(row=row, column=2, value=p.angajat.nume_complet).border = thin_border
            ws.cell(row=row, column=3, value=p.hotel.nume).border = thin_border
            ws.cell(row=row, column=4, value=p.firma_cod or "").border = thin_border
            ws.cell(row=row, column=5, value=p.ore).border = thin_border
            cost = _calc_cost(p)
            total_cost += cost
            ws.cell(row=row, column=6, value=round(cost, 2) if cost else "").border = thin_border
            ws.cell(row=row, column=7, value=p.saptamana or "").border = thin_border

        row += 1
        ws.cell(row=row, column=4, value="TOTAL:").font = Font(bold=True)
        ws.cell(row=row, column=5, value=sum(p.ore for p in pontaje)).font = Font(bold=True)
        if total_cost:
            ws.cell(row=row, column=6, value=round(total_cost, 2)).font = Font(bold=True)
    else:
        headers_map = {"angajat": "Angajat", "hotel": "Hotel", "firma": "Firma"}
        headers = [headers_map.get(grupare, "Grup"), "Total Ore", "Zile Lucrate", "Cost (RON)"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

        grouped = _group_pontaje(pontaje, grupare)
        for key, vals in sorted(grouped.items()):
            row += 1
            ws.cell(row=row, column=1, value=key).border = thin_border
            ws.cell(row=row, column=2, value=round(vals["ore"], 1)).border = thin_border
            ws.cell(row=row, column=3, value=vals["zile"]).border = thin_border
            ws.cell(row=row, column=4, value=round(vals["cost"], 2) if vals["cost"] else "").border = thin_border

        row += 1
        ws.cell(row=row, column=1, value="TOTAL:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=round(sum(v["ore"] for v in grouped.values()), 1)).font = Font(bold=True)
        total_cost = sum(v["cost"] for v in grouped.values())
        if total_cost:
            ws.cell(row=row, column=4, value=round(total_cost, 2)).font = Font(bold=True)

    # Auto-width
    for col in ws.columns:
        lengths = []
        for cell in col:
            if cell.value is not None:
                try:
                    lengths.append(len(str(cell.value)))
                except Exception:
                    pass
        if lengths and hasattr(col[0], "column_letter"):
            ws.column_dimensions[col[0].column_letter].width = min(max(lengths) + 3, 40)

    return wb


def _calc_cost(pontaj):
    """Calculate cost for a single pontaj entry."""
    contract = ContractAngajat.query.filter_by(angajat_id=pontaj.angajat_id).first()
    if contract and contract.tarif_orar:
        return pontaj.ore * contract.tarif_orar
    return 0


def _group_pontaje(pontaje, grupare):
    grouped = {}
    for p in pontaje:
        if grupare == "angajat":
            key = p.angajat.nume_complet
        elif grupare == "hotel":
            key = p.hotel.nume
        else:
            key = p.firma_cod or "N/A"
        if key not in grouped:
            grouped[key] = {"ore": 0, "zile": 0, "cost": 0}
        grouped[key]["ore"] += p.ore
        grouped[key]["zile"] += 1
        grouped[key]["cost"] += _calc_cost(p)
    return grouped


def _filter_pontaje(form):
    """Build filtered pontaj query from form data."""
    query = db.session.query(Pontaj).join(Angajat).join(Hotel)
    angajat_id = form.get("angajat_id", "")
    hotel_id = form.get("hotel_id", "")
    firma_cod = form.get("firma_cod", "")
    data_start = form.get("data_start", "")
    data_end = form.get("data_end", "")

    if angajat_id:
        query = query.filter(Pontaj.angajat_id == int(angajat_id))
    if hotel_id:
        query = query.filter(Pontaj.hotel_id == int(hotel_id))
    if firma_cod:
        query = query.filter(Pontaj.firma_cod == firma_cod)
    if data_start:
        query = query.filter(Pontaj.data >= date.fromisoformat(data_start))
    if data_end:
        query = query.filter(Pontaj.data <= date.fromisoformat(data_end))

    return query.order_by(Pontaj.data, Angajat.nume_complet)


def _get_excluded_pairs():
    """Return a set of frozensets for excluded duplicate pairs."""
    exclusions = DuplicateExclusion.query.all()
    return {frozenset([e.angajat_id_1, e.angajat_id_2]) for e in exclusions}


def _iso_week_start(d):
    """Return Monday of the ISO week containing date d."""
    return d - timedelta(days=d.weekday())


def register_routes(app):

    # --- AUTH ---
    @app.route("/login", methods=["GET", "POST"])
    def login():
        if current_user.is_authenticated:
            return redirect(url_for("index"))
        if request.method == "POST":
            username = request.form.get("username", "").strip()
            password = request.form.get("password", "")
            user = User.query.filter_by(username=username).first()
            if user and user.check_password(password):
                login_user(user, remember=True)
                log_audit("LOGIN", "User", user.id, f"Login: {username}")
                db.session.commit()
                return redirect(request.args.get("next") or url_for("index"))
            flash("Username sau parola incorecta.", "danger")
        return render_template("auth/login.html")

    @app.route("/logout")
    @login_required
    def logout():
        log_audit("LOGOUT", "User", current_user.id)
        db.session.commit()
        logout_user()
        return redirect(url_for("login"))

    @app.route("/register", methods=["GET", "POST"])
    def register():
        if User.query.count() > 0 and not (current_user.is_authenticated and current_user.is_admin):
            flash("Doar adminul poate crea conturi noi.", "danger")
            return redirect(url_for("login"))
        if request.method == "POST":
            username = request.form.get("username", "").strip()
            password = request.form.get("password", "")
            nume = request.form.get("nume_complet", "").strip()
            role = request.form.get("role", "editor")
            # Only admins can set roles; non-admins default to editor
            if not (current_user.is_authenticated and current_user.role == "admin"):
                role = "editor"
            if role not in ("admin", "editor", "viewer"):
                role = "editor"
            if not username or not password:
                flash("Completeaza toate campurile.", "danger")
                return render_template("auth/register.html")
            if User.query.filter_by(username=username).first():
                flash("Username-ul exista deja.", "danger")
                return render_template("auth/register.html")
            user = User(username=username, nume_complet=nume, role=role,
                        is_admin=(role == "admin"))
            user.set_password(password)
            db.session.add(user)
            db.session.commit()
            flash(f"Cont creat: {username} (rol: {role})", "success")
            return redirect(url_for("lista_users") if current_user.is_authenticated else url_for("login"))
        return render_template("auth/register.html")

    @app.route("/users")
    @login_required
    @require_role("admin")
    def lista_users():
        users = User.query.order_by(User.username).all()
        return render_template("auth/users.html", users=users)

    @app.route("/toggle-dark-mode", methods=["POST"])
    @login_required
    def toggle_dark_mode():
        current_user.dark_mode = not current_user.dark_mode
        db.session.commit()
        return jsonify({"dark_mode": current_user.dark_mode})

    # --- DASHBOARD ---
    @app.route("/")
    @login_required
    def index():
        total_angajati = Angajat.query.filter_by(activ=True).count()
        total_pontaje = Pontaj.query.count()
        total_hoteluri = Hotel.query.count()
        total_firme = Firma.query.count()
        recent_pontaje = Pontaj.query.order_by(Pontaj.data.desc()).limit(10).all()

        # Chart data - ore per luna (last 6 months)
        chart_data = []
        today = date.today()
        for i in range(5, -1, -1):
            m = today.month - i
            y = today.year
            while m <= 0:
                m += 12
                y -= 1
            start = date(y, m, 1)
            if m == 12:
                end = date(y + 1, 1, 1) - timedelta(days=1)
            else:
                end = date(y, m + 1, 1) - timedelta(days=1)
            total = db.session.query(db.func.sum(Pontaj.ore)).filter(
                Pontaj.data >= start, Pontaj.data <= end
            ).scalar() or 0
            chart_data.append({
                "label": f"{start.strftime('%b %Y')}",
                "value": round(float(total), 1)
            })

        return render_template(
            "index.html",
            total_angajati=total_angajati,
            total_pontaje=total_pontaje,
            total_hoteluri=total_hoteluri,
            total_firme=total_firme,
            recent_pontaje=recent_pontaje,
            chart_data=chart_data,
        )

    # --- NOTIFICATIONS ---
    @app.route("/notificari")
    @login_required
    def notificari():
        notifs = Notification.query.filter_by(user_id=current_user.id).order_by(
            Notification.created_at.desc()
        ).limit(50).all()
        return render_template("notificari.html", notificari=notifs)

    @app.route("/notificari/citeste-toate", methods=["POST"])
    @login_required
    def citeste_notificari():
        Notification.query.filter_by(user_id=current_user.id, citit=False).update({"citit": True})
        db.session.commit()
        return redirect(url_for("notificari"))

    # --- GLOBAL SEARCH ---
    @app.route("/search")
    @login_required
    def global_search():
        q = request.args.get("q", "").strip()
        if not q or len(q) < 2:
            return jsonify({"results": []})

        results = []
        # Search angajati
        angajati = Angajat.query.filter(
            Angajat.nume_complet.ilike(f"%{q}%")
        ).limit(5).all()
        for a in angajati:
            results.append({
                "type": "angajat", "icon": "person",
                "text": a.nume_complet,
                "url": url_for("edit_angajat", id=a.id),
                "badge": "Activ" if a.activ else "Inactiv"
            })
        # Search hoteluri
        hoteluri = Hotel.query.filter(Hotel.nume.ilike(f"%{q}%")).limit(3).all()
        for h in hoteluri:
            results.append({
                "type": "hotel", "icon": "building",
                "text": h.nume,
                "url": url_for("rapoarte") + f"?hotel_id={h.id}",
            })
        # Search firme
        firme = Firma.query.filter(
            db.or_(Firma.nume.ilike(f"%{q}%"), Firma.cod.ilike(f"%{q}%"))
        ).limit(3).all()
        for f in firme:
            results.append({
                "type": "firma", "icon": "building",
                "text": f"{f.cod} - {f.nume}",
                "url": url_for("edit_firma", id=f.id),
            })

        return jsonify({"results": results})

    # --- ANGAJATI ---
    @app.route("/angajati")
    @login_required
    def lista_angajati():
        show_inactive = request.args.get("inactive", "0") == "1"
        query = Angajat.query
        if not show_inactive:
            query = query.filter_by(activ=True)
        angajati = query.order_by(Angajat.nume_complet).all()
        return render_template("angajati/lista.html", angajati=angajati, show_inactive=show_inactive)

    @app.route("/angajati/nou", methods=["GET", "POST"])
    @login_required
    @require_role("admin", "editor")
    def angajat_nou():
        firme = Firma.query.order_by(Firma.nume).all()
        if request.method == "POST":
            nume = request.form.get("nume", "").strip()
            prenume = request.form.get("prenume", "").strip()
            cnp = request.form.get("cnp", "").strip() or None
            adresa = request.form.get("adresa", "").strip() or None
            telefon = request.form.get("telefon", "").strip() or None
            email = request.form.get("email", "").strip() or None

            if not nume:
                flash("Numele este obligatoriu.", "danger")
                return render_template("angajati/form.html", firme=firme)

            angajat = Angajat(
                nume=nume, prenume=prenume,
                nume_complet=f"{nume} {prenume}".strip(),
                cnp=cnp, adresa=adresa, telefon=telefon, email=email,
            )
            db.session.add(angajat)
            db.session.flush()

            _save_contracts(angajat)
            log_audit("CREATE", "Angajat", angajat.id, f"Creat: {angajat.nume_complet}")
            db.session.commit()
            flash(f"Angajat {angajat.nume_complet} adaugat.", "success")
            return redirect(url_for("lista_angajati"))
        return render_template("angajati/form.html", firme=firme)

    @app.route("/angajati/<int:id>/edit", methods=["GET", "POST"])
    @login_required
    @require_role("admin", "editor")
    def edit_angajat(id):
        angajat = Angajat.query.get_or_404(id)
        firme = Firma.query.order_by(Firma.nume).all()
        if request.method == "POST":
            angajat.nume = request.form.get("nume", "").strip()
            angajat.prenume = request.form.get("prenume", "").strip()
            angajat.nume_complet = f"{angajat.nume} {angajat.prenume}".strip()
            angajat.cnp = request.form.get("cnp", "").strip() or None
            angajat.adresa = request.form.get("adresa", "").strip() or None
            angajat.telefon = request.form.get("telefon", "").strip() or None
            angajat.email = request.form.get("email", "").strip() or None
            angajat.activ = request.form.get("activ") == "on"

            ContractAngajat.query.filter_by(angajat_id=angajat.id).delete()
            _save_contracts(angajat)
            log_audit("UPDATE", "Angajat", angajat.id, f"Editat: {angajat.nume_complet}")
            db.session.commit()
            flash(f"Angajat {angajat.nume_complet} actualizat.", "success")
            return redirect(url_for("lista_angajati"))
        return render_template("angajati/form.html", angajat=angajat, firme=firme)

    @app.route("/angajati/<int:id>/delete", methods=["POST"])
    @login_required
    @require_role("admin", "editor")
    def delete_angajat(id):
        angajat = Angajat.query.get_or_404(id)
        angajat.activ = False
        log_audit("DELETE", "Angajat", angajat.id, f"Dezactivat: {angajat.nume_complet}")
        db.session.commit()
        flash(f"Angajat {angajat.nume_complet} dezactivat.", "warning")
        return redirect(url_for("lista_angajati"))

    # -----------------------------------------------------------------------
    # EMPLOYEE PROFILE (Feature 3)
    # -----------------------------------------------------------------------
    @app.route("/angajati/<int:id>/profil")
    @login_required
    def profil_angajat(id):
        angajat = Angajat.query.get_or_404(id)
        contracte = ContractAngajat.query.filter_by(angajat_id=angajat.id).all()

        # Total hours per month (last 6 months) for chart
        today = date.today()
        monthly_hours = []
        for i in range(5, -1, -1):
            m = today.month - i
            y = today.year
            while m <= 0:
                m += 12
                y -= 1
            start = date(y, m, 1)
            if m == 12:
                end = date(y + 1, 1, 1) - timedelta(days=1)
            else:
                end = date(y, m + 1, 1) - timedelta(days=1)
            total = db.session.query(db.func.sum(Pontaj.ore)).filter(
                Pontaj.angajat_id == angajat.id,
                Pontaj.data >= start, Pontaj.data <= end
            ).scalar() or 0
            monthly_hours.append({
                "label": f"{start.strftime('%b %Y')}",
                "value": round(float(total), 1)
            })

        # Hotel breakdown (pie chart data)
        hotel_breakdown = db.session.query(
            Hotel.nume,
            db.func.sum(Pontaj.ore).label("total_ore")
        ).join(Pontaj).filter(
            Pontaj.angajat_id == angajat.id
        ).group_by(Hotel.nume).all()
        hotel_data = [{"label": h[0], "value": round(float(h[1]), 1)} for h in hotel_breakdown]

        # Paginated pontaj history
        page = request.args.get("page", 1, type=int)
        pontaje = (
            Pontaj.query.filter_by(angajat_id=angajat.id)
            .join(Hotel)
            .order_by(Pontaj.data.desc())
            .paginate(page=page, per_page=30, error_out=False)
        )

        return render_template(
            "angajati/profil.html",
            angajat=angajat,
            contracte=contracte,
            monthly_hours=monthly_hours,
            hotel_data=hotel_data,
            pontaje=pontaje,
        )

    # -----------------------------------------------------------------------
    # DUPLICATE DETECTION & MERGE (updated with "Keep Both" - Feature 1)
    # -----------------------------------------------------------------------
    @app.route("/angajati/duplicate")
    @login_required
    def detectare_duplicate():
        angajati = Angajat.query.filter_by(activ=True).order_by(Angajat.nume_complet).all()
        excluded_pairs = _get_excluded_pairs()
        duplicates = []
        seen = set()
        for i, a1 in enumerate(angajati):
            for a2 in angajati[i + 1:]:
                pair_key = tuple(sorted([a1.id, a2.id]))
                pair_set = frozenset([a1.id, a2.id])
                # Skip if already excluded
                if pair_set in excluded_pairs:
                    continue
                ratio = difflib.SequenceMatcher(
                    None, a1.nume_complet.lower(), a2.nume_complet.lower()
                ).ratio()
                if ratio >= 0.75:
                    if pair_key not in seen:
                        seen.add(pair_key)
                        duplicates.append({
                            "a1": a1, "a2": a2,
                            "similarity": round(ratio * 100)
                        })
        duplicates.sort(key=lambda x: x["similarity"], reverse=True)
        excluded_count = DuplicateExclusion.query.count()
        return render_template("angajati/duplicate.html", duplicates=duplicates, excluded_count=excluded_count)

    @app.route("/angajati/merge", methods=["POST"])
    @login_required
    @require_role("admin", "editor")
    def merge_angajati():
        keep_id = int(request.form.get("keep_id"))
        remove_id = int(request.form.get("remove_id"))
        keep = Angajat.query.get_or_404(keep_id)
        remove = Angajat.query.get_or_404(remove_id)

        # Move pontaje from remove to keep
        Pontaj.query.filter_by(angajat_id=remove.id).update({"angajat_id": keep.id})
        # Move contracts
        for c in remove.contracte:
            existing = ContractAngajat.query.filter_by(
                angajat_id=keep.id, firma_id=c.firma_id
            ).first()
            if not existing:
                c.angajat_id = keep.id
            else:
                db.session.delete(c)
        # Fill missing data
        if not keep.cnp and remove.cnp:
            keep.cnp = remove.cnp
        if not keep.adresa and remove.adresa:
            keep.adresa = remove.adresa
        if not keep.telefon and remove.telefon:
            keep.telefon = remove.telefon

        remove.activ = False
        log_audit("MERGE", "Angajat", keep.id,
                  f"Unificat: {remove.nume_complet} -> {keep.nume_complet}")
        db.session.commit()
        flash(f"'{remove.nume_complet}' a fost unificat in '{keep.nume_complet}'.", "success")
        return redirect(url_for("detectare_duplicate"))

    @app.route("/angajati/exclude-duplicate", methods=["POST"])
    @login_required
    @require_role("admin", "editor")
    def exclude_duplicate():
        """Mark two employees as NOT duplicates (keep both)."""
        id1 = int(request.form.get("id1"))
        id2 = int(request.form.get("id2"))
        # Always store with smaller id first
        a, b = min(id1, id2), max(id1, id2)
        existing = DuplicateExclusion.query.filter_by(
            angajat_id_1=a, angajat_id_2=b
        ).first()
        if not existing:
            excl = DuplicateExclusion(angajat_id_1=a, angajat_id_2=b)
            db.session.add(excl)
            log_audit("EXCLUDE_DUP", "Angajat", a,
                      f"Exclus din duplicate: {a} si {b}")
            db.session.commit()
        flash("Perechea a fost marcata ca non-duplicat.", "info")
        return redirect(url_for("detectare_duplicate"))

    # --- FIRME ---
    @app.route("/firme")
    @login_required
    def lista_firme():
        firme = Firma.query.order_by(Firma.nume).all()
        return render_template("firme/lista.html", firme=firme)

    @app.route("/firme/nou", methods=["GET", "POST"])
    @login_required
    @require_role("admin", "editor")
    def firma_noua():
        if request.method == "POST":
            cod = request.form.get("cod", "").strip().upper()
            nume = request.form.get("nume", "").strip()
            if not cod or not nume:
                flash("Codul si numele sunt obligatorii.", "danger")
                return render_template("firme/form.html")
            firma = Firma(cod=cod, nume=nume)
            db.session.add(firma)
            log_audit("CREATE", "Firma", firma.id, f"Firma noua: {cod} - {nume}")
            db.session.commit()
            flash(f"Firma {nume} adaugata.", "success")
            return redirect(url_for("lista_firme"))
        return render_template("firme/form.html")

    @app.route("/firme/<int:id>/edit", methods=["GET", "POST"])
    @login_required
    @require_role("admin", "editor")
    def edit_firma(id):
        firma = Firma.query.get_or_404(id)
        if request.method == "POST":
            firma.cod = request.form.get("cod", "").strip().upper()
            firma.nume = request.form.get("nume", "").strip()
            log_audit("UPDATE", "Firma", firma.id, f"Editata: {firma.cod}")
            db.session.commit()
            flash(f"Firma {firma.nume} actualizata.", "success")
            return redirect(url_for("lista_firme"))
        return render_template("firme/form.html", firma=firma)

    # --- IMPORT ---
    @app.route("/import", methods=["GET", "POST"])
    @login_required
    @require_role("admin", "editor")
    def import_page():
        if request.method == "POST":
            file = request.files.get("file")
            if not file or not file.filename:
                flash("Selecteaza un fisier.", "danger")
                return redirect(url_for("import_page"))

            content = file.read()
            filename = file.filename.lower()
            try:
                if filename.endswith(".zip"):
                    parsed_list = process_zip_file(content)
                elif filename.endswith(".xlsx"):
                    parsed_list = [parse_excel_file(content, file.filename)]
                else:
                    flash("Format invalid. Acceptam .xlsx sau .zip", "danger")
                    return redirect(url_for("import_page"))
            except Exception as e:
                flash(f"Eroare la citirea fisierului: {e}", "danger")
                return redirect(url_for("import_page"))

            all_new = set()
            for parsed in parsed_list:
                all_new.update(parsed["new_employees"])

            # Store parsed data for preview
            session["pending_import"] = []
            for p in parsed_list:
                session["pending_import"].append({
                    "week_period": p["week_period"],
                    "filename": p["filename"],
                    "entries": p["entries"],
                })

            if all_new:
                session["new_employees"] = sorted(all_new)
                return redirect(url_for("register_new_employees"))

            return redirect(url_for("import_preview"))

        recent_imports = (
            db.session.query(Pontaj.fisier_sursa, db.func.count(Pontaj.id))
            .group_by(Pontaj.fisier_sursa)
            .order_by(Pontaj.fisier_sursa.desc())
            .limit(20).all()
        )
        return render_template("import.html", recent_imports=recent_imports)

    @app.route("/import/preview")
    @login_required
    @require_role("admin", "editor")
    def import_preview():
        pending = session.get("pending_import", [])
        if not pending:
            flash("Nu exista date de importat.", "info")
            return redirect(url_for("import_page"))

        total_entries = sum(len(p["entries"]) for p in pending)
        files = [p["filename"] for p in pending]

        # Collect unique names, hotels, date range
        names = set()
        hotels = set()
        dates = []
        for p in pending:
            for e in p["entries"]:
                names.add(e["name"])
                if e["hotel"]:
                    hotels.add(e["hotel"])

        # ---- Import Validation (Feature 4) ----
        warnings = []
        # Group entries by (name, date) for multi-hotel and >16h checks
        by_name_date = defaultdict(list)
        for p in pending:
            for e in p["entries"]:
                entry_date = e.get("date", "")
                by_name_date[(e["name"], entry_date)].append(e)

        for (name, entry_date), entries in by_name_date.items():
            total_hours = sum(float(en.get("hours", 0) or 0) for en in entries)
            if total_hours > 16:
                warnings.append({
                    "type": "high_hours",
                    "message": f"{name} are {total_hours}h in data {entry_date} (>16h)",
                    "name": name, "date": entry_date,
                })
            hotel_set = {en.get("hotel") for en in entries if en.get("hotel")}
            if len(hotel_set) > 1:
                warnings.append({
                    "type": "multi_hotel",
                    "message": f"{name} apare la {len(hotel_set)} hoteluri in {entry_date}: {', '.join(hotel_set)}",
                    "name": name, "date": entry_date,
                })

        return render_template(
            "import_preview.html",
            total_entries=total_entries,
            total_files=len(files),
            files=files,
            names=sorted(names),
            hotels=sorted(hotels),
            periods=[p["week_period"] for p in pending],
            warnings=warnings,
        )

    @app.route("/import/confirm", methods=["POST"])
    @login_required
    @require_role("admin", "editor")
    def import_confirm():
        pending = session.pop("pending_import", [])
        if not pending:
            flash("Nu exista date de importat.", "info")
            return redirect(url_for("import_page"))

        total_stats = {"imported": 0, "skipped_duplicate": 0, "errors": []}
        for p in pending:
            stats = import_entries(p)
            total_stats["imported"] += stats["imported"]
            total_stats["skipped_duplicate"] += stats["skipped_duplicate"]
            total_stats["errors"].extend(stats["errors"])

        log_audit("IMPORT", "Pontaj", None,
                  f"Import: {total_stats['imported']} noi, {total_stats['skipped_duplicate']} actualizate")
        add_notification(
            f"Import finalizat: {total_stats['imported']} inregistrari noi",
            tip="success", link=url_for("lista_pontaje")
        )
        db.session.commit()

        flash(
            f"Import finalizat: {total_stats['imported']} inregistrari noi, "
            f"{total_stats['skipped_duplicate']} actualizate.", "success"
        )
        if total_stats["errors"]:
            flash(f"Erori: {'; '.join(total_stats['errors'][:5])}", "warning")
        return redirect(url_for("import_page"))

    @app.route("/import/angajati-noi", methods=["GET", "POST"])
    @login_required
    @require_role("admin", "editor")
    def register_new_employees():
        new_employees = session.get("new_employees", [])
        firme = Firma.query.order_by(Firma.nume).all()
        if not new_employees:
            flash("Nu exista angajati noi.", "info")
            return redirect(url_for("import_page"))

        # Find similar names for duplicate warning
        similar_map = {}
        for name in new_employees:
            similar = find_similar_names(name)
            if similar:
                similar_map[name] = similar

        # Feature 6: Fuzzy auto-correct - find names > 85% similar for dropdown
        fuzzy_map = {}
        for name in new_employees:
            matches = find_similar_names(name, threshold=0.85)
            if matches:
                fuzzy_map[name] = matches

        if request.method == "POST":
            mapped_count = 0
            created_count = 0
            for name in new_employees:
                safe_name = name.replace(" ", "_")
                # Check if user chose to map to existing employee
                map_to = request.form.get(f"map_to_{safe_name}", "")
                if map_to and map_to != "new":
                    # User chose to map this name to an existing employee
                    mapped_count += 1
                    # Store the mapping in session so import_entries can use it
                    mappings = session.get("name_mappings", {})
                    mappings[name] = int(map_to)
                    session["name_mappings"] = mappings
                    continue

                cnp = request.form.get(f"cnp_{safe_name}", "").strip() or None
                adresa = request.form.get(f"adresa_{safe_name}", "").strip() or None
                telefon = request.form.get(f"telefon_{safe_name}", "").strip() or None
                firma_id = request.form.get(f"firma_{safe_name}", "")
                nr_contract = request.form.get(f"contract_{safe_name}", "").strip() or None
                functie = request.form.get(f"functie_{safe_name}", "").strip() or None

                angajat = create_angajat_from_name(name)
                angajat.cnp = cnp
                angajat.adresa = adresa
                angajat.telefon = telefon
                created_count += 1

                if firma_id:
                    contract = ContractAngajat(
                        angajat_id=angajat.id, firma_id=int(firma_id),
                        numar_contract=nr_contract, functie=functie,
                    )
                    contract.genereaza_cod()
                    db.session.add(contract)

            details = f"Angajati noi din import: {created_count} creati"
            if mapped_count:
                details += f", {mapped_count} mapati la existenti"
            log_audit("CREATE", "Angajat", None, details)
            add_notification(
                f"{created_count} angajati noi detectati la import",
                tip="warning", link=url_for("lista_angajati")
            )
            db.session.commit()
            session.pop("new_employees", None)
            return redirect(url_for("import_preview"))

        return render_template(
            "import_angajati_noi.html",
            new_employees=new_employees, firme=firme,
            similar_map=similar_map, fuzzy_map=fuzzy_map,
        )

    # --- PONTAJ ---
    @app.route("/pontaj")
    @login_required
    def lista_pontaje():
        page = request.args.get("page", 1, type=int)
        pontaje = (
            Pontaj.query.join(Angajat).join(Hotel)
            .order_by(Pontaj.data.desc(), Angajat.nume_complet)
            .paginate(page=page, per_page=50, error_out=False)
        )
        return render_template("pontaj/lista.html", pontaje=pontaje)

    @app.route("/pontaj/nou", methods=["GET", "POST"])
    @login_required
    @require_role("admin", "editor")
    def pontaj_nou():
        angajati = Angajat.query.filter_by(activ=True).order_by(Angajat.nume_complet).all()
        hoteluri = Hotel.query.order_by(Hotel.nume).all()
        firme = Firma.query.order_by(Firma.nume).all()

        if request.method == "POST":
            angajat_id = request.form.get("angajat_id")
            hotel_id = request.form.get("hotel_id")
            data_str = request.form.get("data")
            ore = request.form.get("ore")
            firma_cod = request.form.get("firma_cod", "")

            if not all([angajat_id, hotel_id, data_str, ore]):
                flash("Completeaza toate campurile obligatorii.", "danger")
                return render_template("pontaj/form.html",
                                       angajati=angajati, hoteluri=hoteluri, firme=firme)

            pontaj = Pontaj(
                angajat_id=int(angajat_id),
                hotel_id=int(hotel_id),
                data=date.fromisoformat(data_str),
                ore=float(ore),
                firma_cod=firma_cod or None,
                fisier_sursa="manual",
            )
            db.session.add(pontaj)
            log_audit("CREATE", "Pontaj", None, f"Pontaj manual: {data_str}")
            db.session.commit()
            flash("Pontaj adaugat.", "success")
            return redirect(url_for("lista_pontaje"))

        return render_template("pontaj/form.html",
                               angajati=angajati, hoteluri=hoteluri, firme=firme)

    # --- RAPOARTE ---
    @app.route("/rapoarte", methods=["GET", "POST"])
    @login_required
    def rapoarte():
        angajati = Angajat.query.order_by(Angajat.nume_complet).all()
        hoteluri = Hotel.query.order_by(Hotel.nume).all()
        firme = Firma.query.order_by(Firma.nume).all()

        pontaje = None
        total_ore = 0
        total_cost = 0
        grouped = {}
        grupare = "detaliat"
        filters = {}
        chart_data = None

        if request.method == "POST":
            grupare = request.form.get("grupare", "detaliat")
            filters = {
                "angajat_id": request.form.get("angajat_id", ""),
                "hotel_id": request.form.get("hotel_id", ""),
                "firma_cod": request.form.get("firma_cod", ""),
                "data_start": request.form.get("data_start", ""),
                "data_end": request.form.get("data_end", ""),
                "grupare": grupare,
            }

            pontaje = _filter_pontaje(request.form).all()
            total_ore = sum(p.ore for p in pontaje)
            total_cost = sum(_calc_cost(p) for p in pontaje)

            if grupare != "detaliat":
                grouped = _group_pontaje(pontaje, grupare)

            # Chart data for results
            if pontaje:
                chart_grouped = {}
                for p in pontaje:
                    key = p.data.strftime("%Y-%m")
                    chart_grouped.setdefault(key, 0)
                    chart_grouped[key] += p.ore
                chart_data = [{"label": k, "value": round(v, 1)}
                              for k, v in sorted(chart_grouped.items())]

        return render_template(
            "rapoarte/rapoarte.html",
            angajati=angajati, hoteluri=hoteluri, firme=firme,
            pontaje=pontaje, total_ore=total_ore, total_cost=total_cost,
            grouped=grouped, grupare=grupare, filters=filters,
            chart_data=chart_data,
        )

    @app.route("/rapoarte/export-excel", methods=["POST"])
    @login_required
    def export_excel():
        grupare = request.form.get("grupare", "detaliat")
        filters = {
            "angajat_id": request.form.get("angajat_id", ""),
            "hotel_id": request.form.get("hotel_id", ""),
            "firma_cod": request.form.get("firma_cod", ""),
            "data_start": request.form.get("data_start", ""),
            "data_end": request.form.get("data_end", ""),
        }
        pontaje = _filter_pontaje(request.form).all()
        wb = build_excel_report(pontaje, filters, grupare)
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        log_audit("EXPORT", "Raport", None, f"Export Excel: {len(pontaje)} randuri")
        db.session.commit()
        return send_file(
            output, as_attachment=True,
            download_name=f"raport_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    @app.route("/rapoarte/export-pdf", methods=["POST"])
    @login_required
    def export_pdf():
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet

        grupare = request.form.get("grupare", "detaliat")
        pontaje = _filter_pontaje(request.form).all()

        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=landscape(A4))
        styles = getSampleStyleSheet()
        elements = []

        elements.append(Paragraph(
            f"Raport Pontaj - {datetime.now().strftime('%d.%m.%Y %H:%M')}", styles["Title"]
        ))
        elements.append(Spacer(1, 20))

        if grupare == "detaliat":
            data = [["Data", "Angajat", "Hotel", "Firma", "Ore", "Cost"]]
            total_ore = 0
            total_cost = 0
            for p in pontaje:
                cost = _calc_cost(p)
                total_ore += p.ore
                total_cost += cost
                data.append([
                    p.data.strftime("%d.%m.%Y"),
                    p.angajat.nume_complet,
                    p.hotel.nume,
                    p.firma_cod or "",
                    str(p.ore),
                    f"{cost:.2f}" if cost else "",
                ])
            data.append(["", "", "", "TOTAL:", str(round(total_ore, 1)),
                          f"{total_cost:.2f}" if total_cost else ""])
        else:
            grouped = _group_pontaje(pontaje, grupare)
            header_map = {"angajat": "Angajat", "hotel": "Hotel", "firma": "Firma"}
            data = [[header_map.get(grupare, "Grup"), "Total Ore", "Zile", "Cost"]]
            for key, vals in sorted(grouped.items()):
                data.append([key, str(round(vals["ore"], 1)), str(vals["zile"]),
                              f"{vals['cost']:.2f}" if vals["cost"] else ""])

        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2B579A")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f0f0f0")]),
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ]))
        elements.append(table)

        doc.build(elements)
        output.seek(0)
        log_audit("EXPORT", "Raport", None, f"Export PDF: {len(pontaje)} randuri")
        db.session.commit()
        return send_file(
            output, as_attachment=True,
            download_name=f"raport_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf",
            mimetype="application/pdf",
        )

    @app.route("/rapoarte/comparatie", methods=["GET", "POST"])
    @login_required
    def comparatie_perioade():
        angajati = Angajat.query.order_by(Angajat.nume_complet).all()
        result = None

        if request.method == "POST":
            p1_start = request.form.get("p1_start")
            p1_end = request.form.get("p1_end")
            p2_start = request.form.get("p2_start")
            p2_end = request.form.get("p2_end")
            angajat_id = request.form.get("angajat_id", "")

            def get_period_data(start, end, ang_id=None):
                q = db.session.query(
                    Angajat.nume_complet,
                    db.func.sum(Pontaj.ore).label("total_ore"),
                    db.func.count(Pontaj.id).label("zile"),
                ).join(Pontaj).filter(
                    Pontaj.data >= date.fromisoformat(start),
                    Pontaj.data <= date.fromisoformat(end),
                ).group_by(Angajat.nume_complet)
                if ang_id:
                    q = q.filter(Pontaj.angajat_id == int(ang_id))
                return {r[0]: {"ore": float(r[1]), "zile": r[2]} for r in q.all()}

            d1 = get_period_data(p1_start, p1_end, angajat_id)
            d2 = get_period_data(p2_start, p2_end, angajat_id)
            all_names = sorted(set(list(d1.keys()) + list(d2.keys())))

            comparison = []
            for name in all_names:
                v1 = d1.get(name, {"ore": 0, "zile": 0})
                v2 = d2.get(name, {"ore": 0, "zile": 0})
                diff = v2["ore"] - v1["ore"]
                comparison.append({
                    "name": name,
                    "p1_ore": v1["ore"], "p1_zile": v1["zile"],
                    "p2_ore": v2["ore"], "p2_zile": v2["zile"],
                    "diff": diff,
                    "pct": round(diff / v1["ore"] * 100, 1) if v1["ore"] else 0,
                })

            result = {
                "comparison": comparison,
                "p1_label": f"{p1_start} - {p1_end}",
                "p2_label": f"{p2_start} - {p2_end}",
                "p1_total": sum(v["ore"] for v in d1.values()),
                "p2_total": sum(v["ore"] for v in d2.values()),
                "filters": {"p1_start": p1_start, "p1_end": p1_end,
                             "p2_start": p2_start, "p2_end": p2_end,
                             "angajat_id": angajat_id},
            }

        return render_template("rapoarte/comparatie.html", angajati=angajati, result=result)

    # -----------------------------------------------------------------------
    # OVERTIME REPORT (Feature 7)
    # -----------------------------------------------------------------------
    @app.route("/rapoarte/ore-suplimentare", methods=["GET", "POST"])
    @login_required
    def overtime_report():
        angajati_list = Angajat.query.filter_by(activ=True).order_by(Angajat.nume_complet).all()
        result = None
        filters = {}

        if request.method == "POST":
            data_start = request.form.get("data_start", "")
            data_end = request.form.get("data_end", "")
            mode = request.form.get("mode", "monthly")  # weekly or monthly
            threshold = float(request.form.get("threshold", 160 if mode == "monthly" else 40))
            filters = {
                "data_start": data_start,
                "data_end": data_end,
                "mode": mode,
                "threshold": threshold,
            }

            q = db.session.query(Pontaj).join(Angajat).filter(Angajat.activ == True)
            if data_start:
                q = q.filter(Pontaj.data >= date.fromisoformat(data_start))
            if data_end:
                q = q.filter(Pontaj.data <= date.fromisoformat(data_end))
            pontaje = q.all()

            # Group by employee and period
            emp_periods = defaultdict(lambda: defaultdict(float))
            for p in pontaje:
                if mode == "weekly":
                    # ISO week key
                    iso_year, iso_week, _ = p.data.isocalendar()
                    period_key = f"{iso_year}-W{iso_week:02d}"
                else:
                    period_key = p.data.strftime("%Y-%m")
                emp_periods[p.angajat_id][period_key] += p.ore

            overtime_data = []
            for ang_id, periods in emp_periods.items():
                ang = Angajat.query.get(ang_id)
                for period, total_ore in sorted(periods.items()):
                    if total_ore > threshold:
                        overtime_data.append({
                            "angajat": ang.nume_complet,
                            "angajat_id": ang.id,
                            "period": period,
                            "total_ore": round(total_ore, 1),
                            "surplus": round(total_ore - threshold, 1),
                            "exceeded": True,
                        })

            overtime_data.sort(key=lambda x: x["surplus"], reverse=True)
            result = overtime_data

        return render_template(
            "rapoarte/ore_suplimentare.html",
            angajati=angajati_list,
            result=result,
            filters=filters,
        )

    # -----------------------------------------------------------------------
    # CALENDAR VIEW (Feature 2)
    # -----------------------------------------------------------------------
    @app.route("/calendar")
    @login_required
    def calendar_view():
        # Determine week start (Monday) from query param or default to current week
        week_start_str = request.args.get("week_start", "")
        if week_start_str:
            try:
                week_start = date.fromisoformat(week_start_str)
                # Snap to Monday
                week_start = _iso_week_start(week_start)
            except ValueError:
                week_start = _iso_week_start(date.today())
        else:
            week_start = _iso_week_start(date.today())

        week_end = week_start + timedelta(days=6)
        days = [week_start + timedelta(days=i) for i in range(7)]

        # Get all pontaje for this week
        pontaje = (
            Pontaj.query.join(Angajat).join(Hotel)
            .filter(Pontaj.data >= week_start, Pontaj.data <= week_end)
            .order_by(Angajat.nume_complet, Pontaj.data)
            .all()
        )

        # Build grid: {angajat_id: {date: [{hotel, ore}]}}
        grid = defaultdict(lambda: defaultdict(list))
        angajati_in_week = {}
        for p in pontaje:
            grid[p.angajat_id][p.data].append({
                "hotel": p.hotel.nume,
                "ore": p.ore,
                "firma": p.firma_cod or "",
            })
            angajati_in_week[p.angajat_id] = p.angajat

        # Sort employees by name
        sorted_angajati = sorted(angajati_in_week.values(), key=lambda a: a.nume_complet)

        prev_week = (week_start - timedelta(days=7)).isoformat()
        next_week = (week_start + timedelta(days=7)).isoformat()

        return render_template(
            "calendar.html",
            week_start=week_start,
            week_end=week_end,
            days=days,
            grid=grid,
            sorted_angajati=sorted_angajati,
            prev_week=prev_week,
            next_week=next_week,
        )

    # -----------------------------------------------------------------------
    # PER-HOTEL DASHBOARD (Feature 8)
    # -----------------------------------------------------------------------
    @app.route("/hotel/<int:id>")
    @login_required
    def hotel_dashboard(id):
        hotel = Hotel.query.get_or_404(id)

        # Total hours at this hotel
        total_ore = db.session.query(db.func.sum(Pontaj.ore)).filter(
            Pontaj.hotel_id == hotel.id
        ).scalar() or 0

        # Number of unique employees
        total_angajati = db.session.query(
            db.func.count(db.distinct(Pontaj.angajat_id))
        ).filter(Pontaj.hotel_id == hotel.id).scalar() or 0

        # Total days worked
        total_zile = db.session.query(db.func.count(Pontaj.id)).filter(
            Pontaj.hotel_id == hotel.id
        ).scalar() or 0

        # Employees who worked here with hours breakdown
        emp_data = db.session.query(
            Angajat.id,
            Angajat.nume_complet,
            db.func.sum(Pontaj.ore).label("total_ore"),
            db.func.count(Pontaj.id).label("zile"),
        ).join(Pontaj).filter(
            Pontaj.hotel_id == hotel.id
        ).group_by(Angajat.id, Angajat.nume_complet).order_by(
            db.func.sum(Pontaj.ore).desc()
        ).all()

        employees = [{
            "id": e[0], "nume": e[1],
            "ore": round(float(e[2]), 1), "zile": e[3]
        } for e in emp_data]

        # Monthly breakdown for chart
        monthly = db.session.query(
            db.func.strftime("%Y-%m", Pontaj.data).label("luna"),
            db.func.sum(Pontaj.ore).label("ore"),
        ).filter(
            Pontaj.hotel_id == hotel.id
        ).group_by("luna").order_by("luna").all()
        chart_data = [{"label": m[0], "value": round(float(m[1]), 1)} for m in monthly]

        # Recent pontaje
        recent = (
            Pontaj.query.filter_by(hotel_id=hotel.id)
            .join(Angajat).order_by(Pontaj.data.desc()).limit(20).all()
        )

        return render_template(
            "hotel_dashboard.html",
            hotel=hotel,
            total_ore=round(float(total_ore), 1),
            total_angajati=total_angajati,
            total_zile=total_zile,
            employees=employees,
            chart_data=chart_data,
            recent=recent,
        )

    @app.route("/hotel/<int:id>/export-excel")
    @login_required
    def hotel_export_excel(id):
        hotel = Hotel.query.get_or_404(id)
        pontaje = (
            Pontaj.query.filter_by(hotel_id=hotel.id)
            .join(Angajat).order_by(Pontaj.data, Angajat.nume_complet).all()
        )
        filters = {"hotel_id": str(hotel.id)}
        wb = build_excel_report(pontaje, filters, "detaliat")
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        log_audit("EXPORT", "Hotel", hotel.id, f"Export Excel hotel: {hotel.nume}")
        db.session.commit()
        safe_name = hotel.nume.replace(" ", "_")
        return send_file(
            output, as_attachment=True,
            download_name=f"hotel_{safe_name}_{datetime.now().strftime('%Y%m%d')}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    # -----------------------------------------------------------------------
    # PLANIFICARE / SCHEDULING (Feature 9)
    # -----------------------------------------------------------------------
    @app.route("/planificare", methods=["GET", "POST"])
    @login_required
    def planificare():
        week_start_str = request.args.get("week_start", "")
        if week_start_str:
            try:
                week_start = date.fromisoformat(week_start_str)
                week_start = _iso_week_start(week_start)
            except ValueError:
                week_start = _iso_week_start(date.today())
        else:
            week_start = _iso_week_start(date.today())

        week_end = week_start + timedelta(days=6)
        days = [week_start + timedelta(days=i) for i in range(7)]

        angajati = Angajat.query.filter_by(activ=True).order_by(Angajat.nume_complet).all()
        hoteluri = Hotel.query.order_by(Hotel.nume).all()
        firme = Firma.query.order_by(Firma.nume).all()

        # Get planned shifts for this week
        planificari = Planificare.query.filter(
            Planificare.data >= week_start,
            Planificare.data <= week_end,
        ).all()

        # Build plan grid: {angajat_id: {date: [{hotel, ore, nota}]}}
        plan_grid = defaultdict(lambda: defaultdict(list))
        for pl in planificari:
            plan_grid[pl.angajat_id][pl.data].append({
                "id": pl.id,
                "hotel": pl.hotel.nume,
                "hotel_id": pl.hotel_id,
                "ore": pl.ore_planificate,
                "nota": pl.nota or "",
                "firma": pl.firma_cod or "",
            })

        # Get actual pontaj for comparison
        pontaje = Pontaj.query.filter(
            Pontaj.data >= week_start, Pontaj.data <= week_end
        ).all()
        actual_grid = defaultdict(lambda: defaultdict(float))
        for p in pontaje:
            actual_grid[p.angajat_id][p.data] += p.ore

        prev_week = (week_start - timedelta(days=7)).isoformat()
        next_week = (week_start + timedelta(days=7)).isoformat()

        return render_template(
            "planificare.html",
            week_start=week_start,
            week_end=week_end,
            days=days,
            angajati=angajati,
            hoteluri=hoteluri,
            firme=firme,
            plan_grid=plan_grid,
            actual_grid=actual_grid,
            prev_week=prev_week,
            next_week=next_week,
        )

    @app.route("/planificare/adauga", methods=["POST"])
    @login_required
    @require_role("admin", "editor")
    def planificare_adauga():
        angajat_id = request.form.get("angajat_id")
        hotel_id = request.form.get("hotel_id")
        data_str = request.form.get("data")
        ore = request.form.get("ore", "8")
        firma_cod = request.form.get("firma_cod", "")
        nota = request.form.get("nota", "").strip()
        week_start = request.form.get("week_start", "")

        if not all([angajat_id, hotel_id, data_str]):
            flash("Completeaza angajat, hotel si data.", "danger")
            return redirect(url_for("planificare", week_start=week_start))

        plan = Planificare(
            angajat_id=int(angajat_id),
            hotel_id=int(hotel_id),
            data=date.fromisoformat(data_str),
            ore_planificate=float(ore),
            firma_cod=firma_cod or None,
            nota=nota or None,
        )
        db.session.add(plan)
        log_audit("CREATE", "Planificare", None,
                  f"Planificare: {data_str} angajat={angajat_id}")
        db.session.commit()
        flash("Schimb planificat adaugat.", "success")
        return redirect(url_for("planificare", week_start=week_start))

    @app.route("/planificare/<int:id>/delete", methods=["POST"])
    @login_required
    @require_role("admin", "editor")
    def planificare_delete(id):
        plan = Planificare.query.get_or_404(id)
        week_start = request.form.get("week_start", "")
        db.session.delete(plan)
        log_audit("DELETE", "Planificare", id, "Planificare stearsa")
        db.session.commit()
        flash("Schimb planificat sters.", "info")
        return redirect(url_for("planificare", week_start=week_start))

    # --- AUDIT LOG ---
    @app.route("/audit")
    @login_required
    @require_role("admin")
    def audit_log():
        page = request.args.get("page", 1, type=int)
        logs = AuditLog.query.order_by(AuditLog.created_at.desc()).paginate(
            page=page, per_page=50, error_out=False
        )
        return render_template("audit.html", logs=logs)

    # --- BACKUP ---
    @app.route("/backup", methods=["POST"])
    @login_required
    @require_role("admin")
    def create_backup():
        db_path = app.config["SQLALCHEMY_DATABASE_URI"].replace("sqlite:///", "")
        if not os.path.exists(db_path):
            flash("Backup disponibil doar cu SQLite.", "warning")
            return redirect(url_for("index"))
        backup_dir = os.path.join(app.instance_path, "backups")
        backup_name = f"backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
        backup_path = os.path.join(backup_dir, backup_name)
        shutil.copy2(db_path, backup_path)
        log_audit("BACKUP", "Database", None, f"Backup: {backup_name}")
        db.session.commit()
        flash(f"Backup creat: {backup_name}", "success")
        return redirect(url_for("index"))

    @app.route("/backup/download")
    @login_required
    @require_role("admin")
    def download_backup():
        db_path = app.config["SQLALCHEMY_DATABASE_URI"].replace("sqlite:///", "")
        if os.path.exists(db_path):
            return send_file(db_path, as_attachment=True,
                             download_name=f"employee_db_{datetime.now().strftime('%Y%m%d')}.db")
        flash("Fisierul bazei de date nu a fost gasit.", "danger")
        return redirect(url_for("index"))

    # --- REST API ---
    @app.route("/api/angajati")
    def api_angajati():
        angajati = Angajat.query.filter_by(activ=True).order_by(Angajat.nume_complet).all()
        return jsonify([{
            "id": a.id, "nume": a.nume_complet, "cnp": a.cnp,
            "telefon": a.telefon, "email": a.email,
            "firme": [{"cod": c.firma.cod, "nume": c.firma.nume, "tarif": c.tarif_orar}
                       for c in a.contracte]
        } for a in angajati])

    @app.route("/api/pontaje")
    def api_pontaje():
        data_start = request.args.get("start")
        data_end = request.args.get("end")
        angajat_id = request.args.get("angajat_id")
        q = Pontaj.query.join(Angajat).join(Hotel)
        if data_start:
            q = q.filter(Pontaj.data >= date.fromisoformat(data_start))
        if data_end:
            q = q.filter(Pontaj.data <= date.fromisoformat(data_end))
        if angajat_id:
            q = q.filter(Pontaj.angajat_id == int(angajat_id))
        pontaje = q.order_by(Pontaj.data).limit(1000).all()
        return jsonify([{
            "id": p.id, "data": p.data.isoformat(), "ore": p.ore,
            "angajat": p.angajat.nume_complet, "hotel": p.hotel.nume,
            "firma": p.firma_cod,
        } for p in pontaje])

    @app.route("/api/stats")
    def api_stats():
        return jsonify({
            "angajati": Angajat.query.filter_by(activ=True).count(),
            "pontaje": Pontaj.query.count(),
            "hoteluri": Hotel.query.count(),
            "firme": Firma.query.count(),
            "total_ore": float(db.session.query(db.func.sum(Pontaj.ore)).scalar() or 0),
        })

    # --- HELPER ---
    def _save_contracts(angajat):
        firma_ids = request.form.getlist("firma_id")
        nr_contracte = request.form.getlist("numar_contract")
        functii = request.form.getlist("functie")
        tarife = request.form.getlist("tarif_orar")
        for i, firma_id in enumerate(firma_ids):
            if firma_id:
                tarif = None
                if i < len(tarife) and tarife[i]:
                    try:
                        tarif = float(tarife[i])
                    except ValueError:
                        pass
                contract = ContractAngajat(
                    angajat_id=angajat.id, firma_id=int(firma_id),
                    numar_contract=nr_contracte[i] if i < len(nr_contracte) else None,
                    functie=functii[i] if i < len(functii) else None,
                    tarif_orar=tarif,
                )
                contract.genereaza_cod()
                db.session.add(contract)


app = create_app()

if __name__ == "__main__":
    app.run(debug=True, port=5000)
