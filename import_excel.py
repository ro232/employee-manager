import os
import re
import io
import zipfile
from datetime import date
from openpyxl import load_workbook
from models import db, Angajat, Hotel, Pontaj, Firma, ContractAngajat


def parse_week_period(value):
    """Parse week period from D1 cell like '2026.03.16-22' -> (year, month)"""
    if not value:
        return None, None
    match = re.match(r"(\d{4})\.(\d{2})\.\d{2}", str(value))
    if match:
        return int(match.group(1)), int(match.group(2))
    return None, None


def parse_date_from_row(date_str, week_period):
    """Parse date like '16.03.' using week period for year context."""
    if not date_str:
        return None
    year, _ = parse_week_period(week_period)
    if not year:
        return None
    match = re.match(r"(\d{1,2})\.(\d{1,2})\.?", str(date_str))
    if match:
        day = int(match.group(1))
        month = int(match.group(2))
        try:
            return date(year, month, day)
        except ValueError:
            return None
    return None


def get_or_create_hotel(nume):
    """Get existing hotel or create new one."""
    hotel = Hotel.query.filter_by(nume=nume).first()
    if not hotel:
        hotel = Hotel(nume=nume)
        db.session.add(hotel)
        db.session.flush()
    return hotel


def find_angajat_by_name(nume_complet):
    """Find employee by full name (case-insensitive)."""
    return Angajat.query.filter(
        db.func.lower(Angajat.nume_complet) == db.func.lower(nume_complet.strip())
    ).first()


def parse_excel_file(file_content, filename):
    """Parse a single Excel file and return structured data.

    Returns:
        dict with keys:
            - week_period: str
            - entries: list of dicts with keys: date_str, name, hotel, firma_cod, hours
            - new_employees: set of names not found in DB
    """
    wb = load_workbook(io.BytesIO(file_content), data_only=True)
    ws = wb.active

    week_period = ws.cell(row=1, column=4).value  # D1
    entries = []
    new_employees = set()

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        date_str = row[1].value   # B
        name = row[2].value       # C
        hotel = row[3].value      # D
        firma_cod = row[4].value  # E
        hours = row[5].value      # F

        if not name or not hours:
            continue

        name = str(name).strip()
        hotel = str(hotel).strip() if hotel else None
        firma_cod = str(firma_cod).strip() if firma_cod else "None"

        try:
            hours = float(hours)
        except (ValueError, TypeError):
            continue

        entries.append({
            "date_str": date_str,
            "name": name,
            "hotel": hotel,
            "firma_cod": firma_cod,
            "hours": hours,
        })

        angajat = find_angajat_by_name(name)
        if not angajat:
            new_employees.add(name)

    return {
        "week_period": str(week_period) if week_period else filename,
        "entries": entries,
        "new_employees": new_employees,
        "filename": filename,
    }


def import_entries(parsed_data):
    """Import parsed entries into the database.

    Returns:
        dict with import stats
    """
    stats = {"imported": 0, "skipped_duplicate": 0, "skipped_no_employee": 0, "errors": []}
    week_period = parsed_data["week_period"]
    filename = parsed_data["filename"]

    for entry in parsed_data["entries"]:
        angajat = find_angajat_by_name(entry["name"])
        if not angajat:
            stats["skipped_no_employee"] += 1
            continue

        parsed_date = parse_date_from_row(entry["date_str"], week_period)
        if not parsed_date:
            stats["errors"].append(f"Data invalida: {entry['date_str']} pentru {entry['name']}")
            continue

        hotel = get_or_create_hotel(entry["hotel"]) if entry["hotel"] else None
        if not hotel:
            stats["errors"].append(f"Hotel lipsa pentru {entry['name']} la {entry['date_str']}")
            continue

        existing = Pontaj.query.filter_by(
            angajat_id=angajat.id,
            data=parsed_date,
            hotel_id=hotel.id,
        ).first()

        if existing:
            existing.ore = entry["hours"]
            existing.firma_cod = entry["firma_cod"]
            stats["skipped_duplicate"] += 1
        else:
            pontaj = Pontaj(
                angajat_id=angajat.id,
                hotel_id=hotel.id,
                data=parsed_date,
                ore=entry["hours"],
                firma_cod=entry["firma_cod"],
                saptamana=week_period,
                fisier_sursa=filename,
            )
            db.session.add(pontaj)
            stats["imported"] += 1

    db.session.commit()
    return stats


def create_angajat_from_name(nume_complet):
    """Create a basic employee record from full name."""
    parts = nume_complet.strip().split(maxsplit=1)
    nume = parts[0] if parts else nume_complet
    prenume = parts[1] if len(parts) > 1 else ""

    angajat = Angajat(
        nume=nume,
        prenume=prenume,
        nume_complet=nume_complet.strip(),
    )
    db.session.add(angajat)
    db.session.flush()
    return angajat


def process_zip_file(zip_content):
    """Process a ZIP file containing multiple Excel files.

    Returns list of parsed data per file.
    """
    results = []
    with zipfile.ZipFile(io.BytesIO(zip_content)) as zf:
        for name in sorted(zf.namelist()):
            if name.endswith(".xlsx") and not name.startswith("__MACOSX"):
                with zf.open(name) as f:
                    content = f.read()
                    basename = os.path.basename(name)
                    parsed = parse_excel_file(content, basename)
                    results.append(parsed)
    return results
